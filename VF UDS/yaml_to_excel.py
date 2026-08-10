# -*- coding: utf-8 -*-
import os
import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def format_hex_list(byte_list):
    """
    辅助函数：将 ['0x02', '0x10'] 或 [0x02, 0x10] 格式化为 clean 字符串 "02 10"
    """
    if not byte_list:
        return ""
    res = []
    for x in byte_list:
        if isinstance(x, str):
            res.append(f"{int(x, 16):02X}")
        elif isinstance(x, int):
            res.append(f"{x:02X}")
    return " ".join(res)


def parse_rcv_nodes(rcv_node_list):
    """
    解析 CAN0 中的 rcvNodeID 并拼接为前置条件字符串
    """
    if not rcv_node_list or not isinstance(rcv_node_list, list):
        return "默认背景报文控制"

    lines = ["默认背景报文节点配置："]
    for node in rcv_node_list:
        if not isinstance(node, dict):
            continue
        node_id = node.get("nodeID", "N/A")
        period = node.get("period", "N/A")
        data_str = format_hex_list(node.get("data", []))
        lines.append(f"• ID: {node_id} | 周期: {period}ms | Data: {data_str}")
    return "\n".join(lines)


def process_single_yaml(yaml_path, ws):
    """
    解析单个 YAML 文件并将提取的用例追加写入 Workbook 的 Sheet 中
    """
    file_name = os.path.basename(yaml_path)
    module_name = os.path.splitext(file_name)[0]

    try:
        with open(yaml_path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
    except Exception as e:
        print(f"⚠️ 警告: 读取或解析 YAML 失败，已跳过。文件: {yaml_path}, 错误: {e}")
        return 0

    if not data or not isinstance(data, dict) or "CAN0" not in data:
        print(f"⚠️ 警告: 文件 {file_name} 格式不符合套件字典规范或为空，已跳过。")
        return 0

    can0_config = data.get("CAN0", {})
    if not isinstance(can0_config, dict):
        return 0

    project_name = can0_config.get("projectName", "VF6.PY")
    rcv_nodes_text = parse_rcv_nodes(can0_config.get("rcvNodeID", []))
    uds_section = can0_config.get("UDS", {})
    if not isinstance(uds_section, dict):
        return 0

    case_count = 0

    # 遍历 PositiveResponse 和 NegativeResponse
    for response_type, type_name in [("PositiveResponse", "正向"), ("NegativeResponse", "反向")]:
        group_data = uds_section.get(response_type, {})
        if not isinstance(group_data, dict):
            continue

        test_cases = group_data.get("TestCase", [])
        if not isinstance(test_cases, list):
            continue

        for case in test_cases:
            if not isinstance(case, dict):
                continue

            case_id = case.get("id", "")
            description = case.get("description", "")
            test_instructions = case.get("testInstructions", {}).get("data", [])
            if not isinstance(test_instructions, list):
                test_instructions = []

            steps_text_list = []
            recvs_text_list = []

            # 逐步骤提取描述
            for idx, step in enumerate(test_instructions, 1):
                if not isinstance(step, dict):
                    continue

                # 场景 1: 发送诊断报文
                if "send" in step:
                    send_str = format_hex_list(step.get("send", []))
                    steps_text_list.append(f"{idx}. 发送: {send_str}")

                    recv_str = format_hex_list(step.get("recv", []))
                    if recv_str:
                        recvs_text_list.append(f"{idx}. 期望接收: {recv_str}")
                    else:
                        recvs_text_list.append(f"{idx}. 期望不响应(无回复)")

                # 场景 2: 修改背景节点 (modifyNode)
                elif "modifyNode" in step or (isinstance(step.get("modifyNode"), dict)):
                    mod = step.get("modifyNode") if isinstance(step.get("modifyNode"), dict) else step
                    node_id = mod.get("nodeID", "")
                    node_data = format_hex_list(mod.get("data", []))
                    steps_text_list.append(f"{idx}. 修改背景节点 {node_id} 数据为: {node_data}")
                    recvs_text_list.append(f"{idx}. 节点数据修改生效")

                # 场景 3: 恢复背景节点 (restoreNodes)
                elif step.get("restoreNodes"):
                    steps_text_list.append(f"{idx}. 恢复所有默认背景节点报文")
                    recvs_text_list.append(f"{idx}. 背景节点恢复默认")

                # 场景 4: 清空背景节点 (clearBackground)
                elif step.get("clearBackground"):
                    steps_text_list.append(f"{idx}. 清空所有背景节点报文")
                    recvs_text_list.append(f"{idx}. 背景节点发送已停止")

            steps_final_str = "\n".join(steps_text_list)
            recvs_final_str = "\n".join(recvs_text_list)

            # 写入单行测试用例数据
            row_data = [
                module_name,  # A: 所属模块（取自 YAML 文件名）
                description,  # B: 用例标题
                rcv_nodes_text,  # C: 前置条件
                steps_final_str,  # D: 步骤
                recvs_final_str,  # E: 各步骤预期结果
                case_id,  # F: 关键词
                "P1",  # G: 优先级
                type_name,  # H: 用例类型
                "系统测试阶段",  # I: 适用阶段
                project_name  # J: 适用项目
            ]
            ws.append(row_data)
            case_count += 1

    return case_count


def convert_yamls_to_single_excel(folder_name="output_yamls", output_excel="UDS_Test_Cases_All.xlsx"):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    yaml_folder = os.path.join(current_dir, folder_name)

    if not os.path.exists(yaml_folder):
        print(f"⚠️ 提示: 文件夹 '{folder_name}' 在当前目录下未找到，尝试直接寻找 '{folder_name}'...")
        yaml_folder = folder_name
        if not os.path.exists(yaml_folder):
            print(f"❌ 错误: 找不到文件夹 '{folder_name}'，请确认文件夹位置。")
            return

    yaml_paths = [
        os.path.join(yaml_folder, f)
        for f in os.listdir(yaml_folder)
        if f.endswith(".yaml") or f.endswith(".yml")
    ]
    yaml_paths.sort()

    if not yaml_paths:
        print(f"❌ 错误: 在 '{yaml_folder}' 文件夹下未找到任何 .yaml 或 .yml 文件！")
        return

    print(f"🔍 在 '{yaml_folder}' 目录下找到 {len(yaml_paths)} 个 YAML 文件，开始处理合并...")

    # 创建合并的 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UDS测试用例汇总"

    # 表头定义
    headers = [
        "所属模块", "用例标题", "前置条件", "步骤",
        "各步骤预期结果", "关键词", "优先级", "用例类型",
        "适用阶段", "适用项目"
    ]
    ws.append(headers)

    total_cases = 0
    for path in yaml_paths:
        count = process_single_yaml(path, ws)
        print(f"  └─ 📄 文件: {os.path.basename(path)} => 解析导出 {count} 条用例")
        total_cases += count

    if total_cases == 0:
        print("⚠️ 警告: 未提取到任何有效测试用例，取消生成 Excel。")
        return

    # ------------------ 样式美化与格式设置 ------------------
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # 商务蓝表头
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="微软雅黑", size=10)

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 美化表头
    for col in range(1, 11):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 美化数据行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
        for col_idx, cell in enumerate(row, 1):
            cell.font = data_font
            cell.border = thin_border

            # C, D, E 列（前置、步骤、预期结果）靠左居中显示并自动换行
            if col_idx in [3, 4, 5]:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            # A, B 列靠左居中
            elif col_idx in [1, 2]:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            # F~J 居中对齐
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 动态固定列宽，确保排版美观
    col_widths = {
        'A': 22,  # 所属模块
        'B': 30,  # 用例标题
        'C': 40,  # 前置条件
        'D': 45,  # 步骤
        'E': 45,  # 各步骤预期结果
        'F': 22,  # 关键词
        'G': 10,  # 优先级
        'H': 12,  # 用例类型
        'I': 15,  # 适用阶段
        'J': 12  # 适用项目
    }

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    output_path = os.path.join(current_dir, output_excel)
    wb.save(output_path)

    print("\n" + "=" * 60)
    print(f"🎉 汇总完成！共处理 {len(yaml_paths)} 个 YAML 文件，生成 {total_cases} 条用例。")
    print(f"📁 导出的 Excel 文件路径: {output_path}")
    print("=" * 60)


if __name__ == "__main__":
    # 执行转换，默认扫描 output_yamls 文件夹，并输出到 UDS_Test_Cases_All.xlsx
    convert_yamls_to_single_excel(folder_name="output_yamls", output_excel="UDS_Test_Cases_All.xlsx")
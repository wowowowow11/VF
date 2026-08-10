"""
****************************************************************************
**  【批量全量版】导出 Excel                                              **
**  1. 使用相对路径，自动定位当前脚本所在的 XY 根目录                      **
**  2. 一次性扫描并生成 XY 下所有子目录 (data15, data30...) 的 Excel 文件      **
**  3. 自动生成多 IP 汇总页 & 细分 Sheet 页 + 图表                         **
****************************************************************************
"""

import os
import json
import re
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

def parse_meta_from_json_filename(filename):
    """从 JSON 文件名中解析元数据"""
    pattern = r"^(.*?)_\[(.*?)\]\s*,\s*ACU_SN[:\-]?\s*(.*?)\s*,\s*(Crash\s*[\w\s]+)\.json"
    match = re.search(pattern, filename, re.IGNORECASE)
    if match:
        base_file = match.group(1).strip()
        time_str = match.group(2).strip()
        acu_sn = match.group(3).strip()
        crash_type = match.group(4).strip()
        return base_file, time_str, acu_sn, crash_type
    else:
        return filename, "", "", "Crash X"

def make_smart_sheet_title(base_file, acu_sn, crash_type):
    """生成符合 Excel 31 字符上限规则的 Sheet 名称"""
    clean_sn = re.sub(r'\s+', '', acu_sn)
    clean_crash = crash_type.strip()

    ip_match = re.search(r"(\d+\.\d+\.\d+\.\d+|\d+)", base_file)
    short_file = ip_match.group(1) if ip_match else base_file[:10]

    candidate = f"{short_file}_{clean_sn}_{clean_crash}"
    candidate = re.sub(r'[\/:*?"<>|]', '_', candidate)

    if len(candidate) <= 31:
        return candidate
    else:
        suffix = f"_{clean_sn[:6]}_{clean_crash}"
        prefix_len = 31 - len(suffix)
        return f"{short_file[:prefix_len]}{suffix}"

def extract_time_key(filename):
    """提取时间戳用于自然按时间顺序排序"""
    match = re.search(r"\[(\d{4}-\d{2}-\d{2}\s+[\d\-]+)\]", filename)
    return match.group(1) if match else filename

def add_line_chart_to_sheet(ws, sheet_title, max_cols, max_rows):
    """
    自动为 Sheet 页创建并插入指定大小的折线图 (高 15cm x 宽 30cm)
    """
    if max_cols < 1 or max_rows < 2:
        return

    # 创建折线图对象
    chart = LineChart()
    chart.title = f"{sheet_title} - 波形对比图"
    chart.style = 10  # 经典折线图样式
    chart.y_axis.title = "加速度 (g)"
    chart.x_axis.title = "采样点 (Points)"

    # 引用数据源范围：包含 Row 1 (表头/文件名) 到 Row max_rows (500行数据)
    data = Reference(ws, min_col=1, min_row=1, max_col=max_cols, max_row=max_rows)

    # titles_from_data=True 表示第一行为每条曲线的标签 (JSON文件名)
    chart.add_data(data, titles_from_data=True)

    # 设置图表尺寸 (单位：厘米 cm)
    chart.height = 15.00  # 高度 15.00 厘米
    chart.width = 30.00   # 宽度 30.00 厘米

    # 图表摆放位置：放在数据列最右侧空处
    chart_col_letter = get_column_letter(max_cols + 2)
    ws.add_chart(chart, f"{chart_col_letter}2")

def write_columns_and_chart_to_sheet(ws, sheet_title, file_list, target_dir, header_font, header_fill, header_align):
    """向指定 Sheet 中按列写入数据，并自动插入曲线图"""
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 40

    col_idx = 1
    max_data_rows = 1

    for jf in file_list:
        json_path = os.path.join(target_dir, jf)

        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                content = json.load(f)

            array_vars = []
            for k, v in content.items():
                if isinstance(v, list) and len(v) > 0:
                    array_vars.append((k, v))

            if not array_vars:
                continue

            for var_name, arr_values in array_vars:
                header_title = jf if len(array_vars) == 1 else f"{jf}\n({var_name})"

                # 写入表头
                cell_header = ws.cell(row=1, column=col_idx, value=header_title)
                cell_header.font = header_font
                cell_header.fill = header_fill
                cell_header.alignment = header_align

                # 写入波形数值
                for row_offset, val in enumerate(arr_values, start=2):
                    clean_val = val if val is not None else ""
                    ws.cell(row=row_offset, column=col_idx, value=clean_val)

                if (len(arr_values) + 1) > max_data_rows:
                    max_data_rows = len(arr_values) + 1

                # 设置列宽
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 28
                col_idx += 1

        except Exception as e:
            print(f"  └─ 读取文件失败 [{jf}]: {str(e)}")

    written_cols = col_idx - 1

    # 数据写入完成后，自动生成指定尺寸的折线图
    if written_cols >= 1 and max_data_rows >= 2:
        add_line_chart_to_sheet(ws, sheet_title, max_cols=written_cols, max_rows=max_data_rows)

    return written_cols

def export_vertical_waveforms_to_excel(target_dir, output_excel_path):
    """处理单个文件夹，生成 Excel"""
    if not os.path.exists(target_dir):
        print(f"错误: 目标目录不存在 -> {target_dir}")
        return

    json_files = [f for f in os.listdir(target_dir) if f.lower().endswith('.json')]
    if not json_files:
        print(f"  └─ 在目录 [{target_dir}] 中未找到任何 .json 文件，跳过。")
        return

    json_files.sort(key=extract_time_key)
    folder_name = os.path.basename(os.path.normpath(target_dir))

    # 识别不同 IP 及其最早一条数据 + 细分分组
    earliest_x_per_ip = {}
    earliest_y_per_ip = {}
    sheets_files_map = {}

    for jf in json_files:
        base_file, time_str, acu_sn, crash_type = parse_meta_from_json_filename(jf)

        if "Crash X" in crash_type:
            if base_file not in earliest_x_per_ip:
                earliest_x_per_ip[base_file] = jf
        elif "Crash Y" in crash_type:
            if base_file not in earliest_y_per_ip:
                earliest_y_per_ip[base_file] = jf

        sheet_title = make_smart_sheet_title(base_file, acu_sn, crash_type)
        if sheet_title not in sheets_files_map:
            sheets_files_map[sheet_title] = []
        sheets_files_map[sheet_title].append(jf)

    # 创建 Excel 工作簿及样式
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 移除默认 Sheet

    header_font = Font(name="微软雅黑", size=9, bold=True, color="1F497D")
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    summary_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    total_cols = 0

    # 1. 生成汇总图表 Sheet 1: 各 IP 第一条 Crash X
    sheet_name_summary_x = f"{folder_name}_各IP第一条_X"[:31]
    ws_summary_x = wb.create_sheet(title=sheet_name_summary_x)

    sorted_ips = sorted(earliest_x_per_ip.keys())
    first_x_files = [earliest_x_per_ip[ip] for ip in sorted_ips]
    cols_x = write_columns_and_chart_to_sheet(ws_summary_x, sheet_name_summary_x, first_x_files, target_dir, header_font, summary_fill, header_align)
    total_cols += cols_x

    # 2. 生成汇总图表 Sheet 2: 各 IP 第一条 Crash Y
    sheet_name_summary_y = f"{folder_name}_各IP第一条_Y"[:31]
    ws_summary_y = wb.create_sheet(title=sheet_name_summary_y)

    sorted_ips_y = sorted(earliest_y_per_ip.keys())
    first_y_files = [earliest_y_per_ip[ip] for ip in sorted_ips_y]
    cols_y = write_columns_and_chart_to_sheet(ws_summary_y, sheet_name_summary_y, first_y_files, target_dir, header_font, summary_fill, header_align)
    total_cols += cols_y

    # 3. 生成其余各个 IP 细分的完整 Sheet 页与折线图
    for sheet_name, files_list in sheets_files_map.items():
        ws = wb.create_sheet(title=sheet_name)
        cols = write_columns_and_chart_to_sheet(ws, sheet_name, files_list, target_dir, header_font, header_fill, header_align)
        total_cols += cols

    # 4. 保存文件 (带文件占用提示保护)
    try:
        wb.save(output_excel_path)
        print(f"  └─ 成功生成 Excel: {output_excel_path} (含 {len(wb.sheetnames)} 个 Sheet 页)")
    except PermissionError:
        timestamp = datetime.datetime.now().strftime("%H%M%S")
        backup_excel_path = os.path.join(target_dir, f"{folder_name}_waveform_vertical_{timestamp}.xlsx")
        print(f"  └─ [⚠️ 警告] 主文件正在被 WPS/Excel 打开占用，已存至备用文件: {backup_excel_path}")
        wb.save(backup_excel_path)

def extract_data_num(path):
    """提取自然数字，用于升序排列 data15, data30, data45..."""
    match = re.search(r'data(\d+)', path, re.IGNORECASE)
    return int(match.group(1)) if match else 999

def process_all_json_folders_to_excel(xy_root_dir):
    """一次性递归搜寻 XY 目录下所有的 dataXX 子文件夹并生成 Excel"""
    if not os.path.exists(xy_root_dir):
        print(f"错误: 根目录不存在 -> {xy_root_dir}")
        return

    # 扫描所有包含 .json 文件的子文件夹
    target_subdirs = []
    for root, dirs, files in os.walk(xy_root_dir):
        if any(f.lower().endswith('.json') for f in files):
            target_subdirs.append(root)

    if not target_subdirs:
        print(f"在目录 {xy_root_dir} 及其子目录下未找到任何包含 .json 的文件夹！")
        return

    # 按 data15, data30, data45... 自然数字顺序排序
    target_subdirs.sort(key=extract_data_num)

    print("="*60)
    print(f"【批量生成 Excel】找到 {len(target_subdirs)} 个包含 JSON 的数据文件夹，开始全量处理...")
    print("="*60)

    for sdir in target_subdirs:
        folder_name = os.path.basename(os.path.normpath(sdir))
        output_excel = os.path.join(sdir, f"{folder_name}_waveform_vertical.xlsx")
        print(f"\n正在处理目录: [{folder_name}] ...")
        export_vertical_waveforms_to_excel(sdir, output_excel)

    print("\n" + "="*60)
    print("🎉 所有子目录的 Excel 文件已全部生成完成！")
    print("="*60)

if __name__ == "__main__":
    # 【相对路径】：自动获取当前脚本 json_to_excel.py 所在的 XY 根目录
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

    process_all_json_folders_to_excel(SCRIPT_DIR)